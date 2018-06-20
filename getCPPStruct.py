# -*- coding: utf-8 -*-
"""
Created on Fri May 25 10:51:24 2018

@author: 12600771
"""

import openpyxl

def convertToVar():

    typeDict = {"bool":"BOOL", "double":"LREAL", "float": "REAL", \
                "int":"DINT", "unsigned int":"DWORD"}

    fileName = "InterfaceSymbol.xlsx"
    workbook = openpyxl.load_workbook(fileName)
    sheet1 = workbook["Sheet1"]
    # cell = sheet1.cell(row=3, column=4)
    # print(sheet1.max_row)


    structString = ""
    newStructTypeName = "g_ChannelDataFull"
    newStructName = "g_ChannelDataFullName"
    structBeginString = "TYPE" + "\n" + "\t" + newStructName \
                        + " : STRUCT" + "\n"
    structMainString = ""
    for i in range(2, sheet1.max_row + 1):
#        cellForLongName = sheet1.cell(row = i, column = 4)
#        longName = cellForLongName.value
#        cellForType = sheet1.cell(row = i, column = 2)
#        typeName = cellForType.value
#        cellForShortName = sheet1.cell(row = i, column = 5)
#        shortName = cellForShortName.value
#        name = longName if len(longName) < 32 else shortName
#        if "[" not in name:
#            structMainString += "\t" * 2 + name + " : " \
#                                + typeDict.get(typeName)  + ";\n"
#        else:
#            index1 = name.find("[")
#            index2 = name.find("]")
#            num = name[index1+1: index2]
#            structMainString += "\t" * 2 + name[:index1] + " : " \
#                                + "ARRAY[0.." + num + "] OF " \
#                                + typeDict.get(typeName) + ";\n"
        
        cellForUsage = sheet1.cell(row = i, column = 2)
        usage = cellForUsage.value
        cellForName = sheet1.cell(row = i, column = 8)
        name = cellForName.value
        cellForType = sheet1.cell(row = i, column = 10)
        typeName = cellForType.value
        cellForCheckName = sheet1.cell(row = i, column = 11)
        checkName = cellForCheckName.value
        used = (name == checkName)
        
        if used:
            structMainString += "\t" * 2 + name + " : " + typeName + ";\n"
    
    # sheet 2
    sheet2 = workbook["Sheet2"]
    for i in range(2, sheet2.max_row + 1):
        cellForName = sheet2.cell(row = i, column = 2)
        name = cellForName.value
        cellForType = sheet2.cell(row = i, column = 3)
        typeName = cellForType.value
        structMainString += "\t" * 2 + name + " : " + typeName + ";\n" 

    structEndString = "\t" + "END_STRUCT;\n" + "END_TYPE"
    structString = structBeginString + structMainString + structEndString


    with open("Types.typ", 'w') as f:
        f.write(structString)

    #connect struct name to struct type
    connectionLanguage = "VAR\n\t" + newStructTypeName + " : " \
                        + newStructName + ";\n" + "END_VAR"
    with open("Variables.var", 'w') as f:
        f.write(connectionLanguage)

    