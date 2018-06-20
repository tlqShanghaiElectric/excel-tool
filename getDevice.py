# -*- coding: utf-8 -*-
"""
Created on Mon May 28 08:43:33 2018

@author: 12600771
"""

import openpyxl
import xml.etree.ElementTree as ET

def convertToDevice():
    def indent(elem, level=0):
        i = "\n" + level*"  "
        j = "\n" + (level-1)*"  "
        if len(elem):
            if not elem.text or not elem.text.strip():
                elem.text = i + "  "
            if not elem.tail or not elem.tail.strip():
                elem.tail = i
            for subelem in elem:
                indent(subelem, level+1)
            if not elem.tail or not elem.tail.strip():
                elem.tail = j
        else:
            if level and (not elem.tail or not elem.tail.strip()):
                elem.tail = j
        return elem

    typeDict = {"BOOL":"Boolean", "double":"Double", "REAL": "Single", \
                "DINT":"Int32", "DWORD":"UInt32"}

    fileName = "InterfaceSymbol.xlsx"
    workbook = openpyxl.load_workbook(fileName)
    sheet1 = workbook["Sheet1"]
    sheet2 = workbook["Sheet2"]

    ns_xsi = "http://www.w3.org/2001/XMLSchema-instance"
    ns_xsd = "http://www.w3.org/2001/XMLSchema"
    device_attrib = {"xmlns:xsi":ns_xsi,"xmlns:xsd":ns_xsd}
    device = ET.Element("Device",attrib=device_attrib)
    # print(ET.tostring(root))


    # Create Protocol
# =============================================================================
#     # <Device>
#     #     <Protocol>
# =============================================================================
    protocol = ET.SubElement(device, 'Protocol', attrib={"xsi:type":"PVIProtocol"})

    # Create Variables
# =============================================================================
#     # <Device>
#     #   <Protocol>
#     #       <Variables>
# =============================================================================
    variables = ET.SubElement(protocol, 'Variables')

# =============================================================================
#     # Create ProtocolVariable
#     # <Variables>
#     #   <ProtocolVariable>
#     #   <ProtocolVariable>
# =============================================================================
    for i in range(2, sheet1.max_row+1):
        # parse from excel to text
#        num = 1
#        cellForLongName = sheet1.cell(row = i, column = 4)
#        longName = cellForLongName.value
#        cellForType = sheet1.cell(row = i, column = 2)
#        typeName = cellForType.value
#        cellForShortName = sheet1.cell(row = i, column = 5)
#        shortName = cellForShortName.value
#        name = longName if len(longName) < 32 else shortName
#
#        if "[" in longName:
#            index1 = longName.find("[")
#            index2 = longName.find("]")
#            num = int(longName[index1+1:index2])
#
#        if "[" in name:
#            index1 = name.find("[")
#            name = name[:index1]
        
        cellForName = sheet1.cell(row = i, column = 3)
        name = cellForName.value
        cellForDescription = sheet1.cell(row = i, column = 4)
        description = cellForDescription.value
        cellForType = sheet1.cell(row = i, column = 5)
        typeName = cellForType.value
        cellForInitial = sheet1.cell(row = i, column = 6)
        initial = cellForInitial.value
        cellForRemote = sheet1.cell(row = i, column = 8)
        remoteName = cellForRemote.value
        cellForCheck = sheet1.cell(row = i, column = 11)
        checkName = cellForCheck.value

        # text to xml
# =============================================================================
#         """       
#         <ProtocolVariable xsi:type="PVIProtocolVariable">
#             <Name>P_DisconStateMachineStartTime</Name>
#             <Description>Delay in discon state machine start up</Description>
#             <VariableType>Single</VariableType>
#             <ArraySize>1</ArraySize>
#             <DefaultVal>10</DefaultVal>
#             <MinVal />
#             <MaxVal />
#             <OutgoingScalingFactor />
#             <Unit />
#             <DisplayNames />
#             <RemoteName>P_DisStaMacStaTim</RemoteName>
#         </ProtocolVariable> 
#         """
# =============================================================================

        # Do not parse array for the present
#        if "[" not in name:
        if (remoteName == checkName):
            protocolVariable = ET.SubElement(variables, 'ProtocolVariable', \
                                    attrib={"xsi:type":"PVIProtocolVariable"})
            ET.SubElement(protocolVariable, "Name").text = name
            ET.SubElement(protocolVariable, "Description").text = description
            ET.SubElement(protocolVariable, "VariableType").text = typeName
            ET.SubElement(protocolVariable, "ArraySize").text = "1"
            ET.SubElement(protocolVariable, "DefaultVal").text = str(initial)
            ET.SubElement(protocolVariable, "MinVal")
            ET.SubElement(protocolVariable, "MaxVal")
            ET.SubElement(protocolVariable, "OutgoingScalingFactor")
            ET.SubElement(protocolVariable, "Unit")
            ET.SubElement(protocolVariable, "DisplayNames")
            ET.SubElement(protocolVariable, "RemoteName").text = remoteName

    # sheet2
    for i in range(2, sheet2.max_row+1):
        # parse from excel to text
        cellForName = sheet2.cell(row = i, column = 2)
        name = cellForName.value
        cellForType = sheet2.cell(row = i, column = 3)
        typeName = cellForType.value
        cellForDefaultValue = sheet2.cell(row = i, column = 4)
        defaultValue = cellForDefaultValue.value
        
        # text to xml
        if (remoteName == checkName):
            protocolVariable = ET.SubElement(variables, 'ProtocolVariable', \
                                    attrib={"xsi:type":"PVIProtocolVariable"})
            ET.SubElement(protocolVariable, "Name").text = name
            ET.SubElement(protocolVariable, "Description")
            ET.SubElement(protocolVariable, "VariableType")\
                                .text = typeDict.get(typeName)
            ET.SubElement(protocolVariable, "ArraySize").text = "1"
            ET.SubElement(protocolVariable, "DefaultVal").text = str(defaultValue)
            ET.SubElement(protocolVariable, "MinVal")
            ET.SubElement(protocolVariable, "MaxVal")
            ET.SubElement(protocolVariable, "OutgoingScalingFactor")
            ET.SubElement(protocolVariable, "Unit")
            ET.SubElement(protocolVariable, "DisplayNames")
            ET.SubElement(protocolVariable, "RemoteName").text = name

    # Add some other node
# =============================================================================
#     """
#     <Protocol>
#         <Variables>
#         </Variables>
#         <RemoteIP>172.160.22.12</RemoteIP>
#         <RemotePort>11169</RemotePort>
#         <StructName>g_ChannelDataFull</StructName>
#     """
# =============================================================================
    newStructTypeName = "g_ChannelDataFull"
    ET.SubElement(protocol, "RemoteIP").text = "172.160.22.12"
    ET.SubElement(protocol, "RemotePort").text = "11169"
    ET.SubElement(protocol, "StructName").text = newStructTypeName

# =============================================================================
#      <Device>
#          <Name>Turbine Controller</Name>
#          <Description>Symbols for the B&amp;R controller</Description>
#          <ReadFrequencyHz>50</ReadFrequencyHz>
#          <DataStorageConfig>
#              <SourceFile>C:\Users\12600771\Desktop\device.device</SourceFile>
#              <LastLoadedAt>2018-05-28T08:22:51.4819742+08:00</LastLoadedAt>
#              <SourceFileLastModifiedAt>0001-01-01T00:00:00</SourceFileLastModifiedAt>
#              <UpdateType>PromptForReload</UpdateType>
#          </DataStorageConfig>
# =============================================================================

    ET.SubElement(device, "Name").text = "Turbine Controller"
    ET.SubElement(device, "Description").text \
        = "Symbols for the B&amp;R controller"
    ET.SubElement(device, "ReadFrequencyHz").text = str(50)
    dataStorageConfig  = ET.SubElement(device, "DataStorageConfig")
    ET.SubElement(dataStorageConfig, "SourceFile").text \
        = r"C:\Users\12600771\Desktop\device.device"
    ET.SubElement(dataStorageConfig, "LastLoadedAt").text\
        = "2018-05-28T08:22:51.4819742+08:00"
    ET.SubElement(dataStorageConfig, "LastLoadedAt").text \
        = "0001-01-01T00:00:00"
    ET.SubElement(dataStorageConfig, "UpdateType").text = "PromptForReload"

    # Add deviceConfig
    """ 
    <Device>
        <DeviceConfig>
            <Name>Turbine Controller</Name>
            <Assignments />
            <VariableConfig>
            </VariableConfig>
        </DeviceConfig>
    """
    deviceConfig = ET.SubElement(device, "DeviceConfig")
    ET.SubElement(deviceConfig, "Name").text = "Turbine Controller"
    ET.SubElement(deviceConfig, "Assignments")
    variableConfig = ET.SubElement(deviceConfig, "VariableConfig")

# =============================================================================
#     # add protocolVariableConfig
#     """ <VariableConfig>
#         <ProtocolVariableConfig>
#             <Name>CI_AlgSpeedControlSpeedSetpoint</Name>
#             <Usage>ReadContinuously</Usage>
#             <Record>true</Record>
#         </ProtocolVariableConfig>
#         ...
#     </VariableConfig> """
# =============================================================================
    for i in range(2, sheet1.max_row+1):
        # parse from excel to text
        cellForRecord = sheet1.cell(row = i, column = 1)
        record = cellForRecord.value
        cellForUsage = sheet1.cell(row = i, column = 2)
        usage = cellForUsage.value
        cellForName = sheet1.cell(row = i, column = 3)
        name = cellForName.value
        cellForRemote = sheet1.cell(row = i, column = 8)
        remoteName = cellForRemote.value
        cellForCheck = sheet1.cell(row = i, column = 11)
        checkName = cellForCheck.value

        if remoteName == checkName :
            protocolVariableConfig = ET.SubElement(variableConfig, \
                                        "ProtocolVariableConfig")
            ET.SubElement(protocolVariableConfig, "Name").text = name
            ET.SubElement(protocolVariableConfig, "Usage").text = usage
            ET.SubElement(protocolVariableConfig, "Record").text = str(record)

    # sheet2
    for i in range(2, sheet2.max_row+1):     
        # parse from excel to text
        record = "TRUE"
        cellForUsage = sheet2.cell(row = i, column = 1)
        usage = cellForUsage.value
        cellForName = sheet2.cell(row = i, column = 2)
        name = cellForName.value
        
        # text to xml
        protocolVariableConfig = ET.SubElement(variableConfig, \
                                            "ProtocolVariableConfig")
        ET.SubElement(protocolVariableConfig, "Name").text = name
        ET.SubElement(protocolVariableConfig, "Usage").text = usage
        ET.SubElement(protocolVariableConfig, "Record").text = str(record)

    # add more nodes
# =============================================================================
#     """ <Device>
#         <DependentDeviceNames />
#         <Assignments />
#         <Scripts>
#             <ExecutableScript>
#                 <Expression>
#                 </Expression>
#             </ExecutableScript>
#         </Scripts>
#     </Device> """
# =============================================================================
    ET.SubElement(device, "DependentDeviceNames")
    ET.SubElement(device, "Assignments")
    scripts = ET.SubElement(device, "Scripts")
    executableScript = ET.SubElement(scripts, "ExecutableScript")
    ET.SubElement(executableScript, "Expression")

    deviceFile = "device.device"
    indent(device)
    # ET.dump(device)
    tree = ET.ElementTree(device)
    tree.write(deviceFile, encoding="UTF-8", xml_declaration=True)
    # with open(deviceFile, 'w') as f:
    #     f.write(ET.tostring(device,encoding="UTF-8"))
    # ET.tostring(device)