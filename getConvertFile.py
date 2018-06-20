# -*- coding: utf-8 -*-
"""
Created on Fri May 25 16:11:58 2018

@author: 12600771
"""

import openpyxl

def convertToSt():
    typeDict = {"bool":"BOOL", "double":"LREAL", "float": "REAL", "int":"DINT", "unsigned int":"DWORD"}

    convertLanguage = ""
    convertPrefixLanguage = "PROGRAM _CYCLE\n"
    newStructName = "g_ChannelDataFull"
    oldStructName = "g_ChannelData"
    convertMainLanguage = ""
    convertSuffixLanguage = "END_PROGRAM"

    # Parse excel
    fileName = "InterfaceSymbol.xlsx"
    workbook = openpyxl.load_workbook(fileName)
    
    # sheet1
    sheet1 = workbook["Sheet1"]
    for i in range(2, sheet1.max_row+1):
#        cellForLongName = sheet1.cell(row = i, column = 4)
#        longName = cellForLongName.value
#        cellForType = sheet1.cell(row = i, column = 2)
#        typeName = cellForType.value
#        cellForShortName = sheet1.cell(row = i, column = 5)
#        shortName = cellForShortName.value
#        name = longName if len(longName) < 32 else shortName
#        if "[" not in name:
#            convertMainLanguage += "\t" + newStructName + "." + name \
#                                    + " := " + oldStructName + "." + shortName + ";\n"
#        else:
#            index1 = name.find("[")
#            index2 = name.find("]")
#            num = int(name[index1+1: index2])
#            name = name[:index1]
#            shortNameIndex1 = shortName.find("[")
#            shortName = shortName[:shortNameIndex1]
#            for j in range(num+1):     
#                convertMainLanguage += "\t" + newStructName + "." + name \
#                                    + "[" + str(j) + "]" + " := " + oldStructName \
#                                    + "." + shortName + "[" + str(j) + "]" + ";\n"
        cellForUsage = sheet1.cell(row = i, column = 2)
        usage = cellForUsage.value
        cellForName = sheet1.cell(row = i, column = 8)
        name = cellForName.value
#        cellForType = sheet1.cell(row = i, column = 10)
#        typeName = cellForType.value
        cellForCheckName = sheet1.cell(row = i, column = 11)
        checkName = cellForCheckName.value
        used = (name == checkName)
        
        if used:
            convertMainLanguage += "\t" + newStructName + "." + name \
                                + " := " + oldStructName + "." + name + ";\n"
                                
    # sheet2
    sheet2 = workbook["Sheet2"]
    for i in range(2, sheet2.max_row + 1):
        cellForName = sheet2.cell(row = i, column = 2)
        name = cellForName.value
        
        convertMainLanguage += "\t" + newStructName + "." + name \
                                + " := " + name + ";\n"
    

    convertLanguage = convertPrefixLanguage + convertMainLanguage \
                        + convertSuffixLanguage
                        

    # To the file
    with open("Cyclic.st", 'w') as f:
        f.write(convertLanguage)